from django.utils.timezone import now
from django.shortcuts import render, redirect
from django.contrib import messages
import openpyxl
from openpyxl.utils import get_column_letter
from .models import Candidate, Voter, Position
from django.views.decorators.csrf import csrf_exempt,csrf_protect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse


@csrf_exempt  # Ideally, use proper CSRF protection
def voter_login(request):
    if request.method == 'POST':
        unique_id = request.POST.get('unique_id')
        voter_code = request.POST.get('password')
        
        # Authentication logic here
        try:
            voter = Voter.objects.get(unique_id=unique_id, voter_code=voter_code)
            if voter.has_voted:
                messages.info(request, "You have already cast your vote.")
                return redirect('voter_login')
                
            # Store voter ID in session
            request.session['voter_id'] = voter.id
            
            # Redirect to candidate page
            return redirect('candidate_page')  
        except Voter.DoesNotExist:
            messages.error(request, "Invalid login details.")
            return render(request, 'voter_login.html')
    return render(request, 'voter_login.html')

@csrf_exempt
def candidate_page(request):
    if 'voter_id' not in request.session:
        return redirect('voter_login')  # Redirect if not logged in

    if request.method == 'POST':
        selected_ids = request.POST.get('candidates')
        if selected_ids:
            try:
                candidate_ids = [int(id.strip()) for id in selected_ids.split(',')]
            except ValueError:
                messages.error(request, "Invalid candidate IDs.")
                return redirect('candidate_page')

            voted_candidates = []
            for candidate_id in candidate_ids:
                try:
                    candidate = Candidate.objects.get(id=candidate_id)
                    candidate.vote_count += 1
                    candidate.save()
                    voted_candidates.append(candidate.name)
                except Candidate.DoesNotExist:
                    messages.error(request, f"Candidate with ID {candidate_id} not found.")
                    continue

            try:
                voter = Voter.objects.get(id=request.session['voter_id'])
                voter.has_voted = True
                voter.date_voted = now()
                voter.save()
            except Voter.DoesNotExist:
                messages.error(request, "Voter not found.")
                return redirect('voter_login')
            if voter.has_voted:
                messages.info(request, 'You have already cast your vote.')
                return redirect('candidate_page')
                
            if voted_candidates:
                voted_names = ', '.join(voted_candidates)
                messages.success(request, f"You have cast  your vote successfully!")
            else:
                messages.warning(request, "No valid votes were cast.")

            return redirect('candidate_page')

        else:
            messages.error(request, "No candidates were selected.")
            return redirect('candidate_page')

    candidates = Candidate.objects.all()
    return render(request, 'candidate_page.html', {'candidates': candidates})


@csrf_exempt  # Ideally, remove this and use CSRF tokens
def voter_logout(request):
    request.session.flush()  # Clear all session data (logs out the voter)
    return redirect('voter_login')


# Results view
def results_view(request):
    positions = Position.objects.all()
    results = {}

    # Calculate vote counts for each candidate under each position
    for position in positions:
        candidates = Candidate.objects.filter(position=position)
        position_result = []
        
        for candidate in candidates:
            position_result.append({
                'candidate':candidate,
                'vote_count':candidate.vote_count
            })
        # Sort candidates by the highest vote count
        sorted_results = sorted(position_result, key=lambda x:x['vote_count'], reverse=True)
        results[position] = sorted_results
    return render(request, 'results.html', {'results': results})

# Chart view
def chart_view(request):
    candidates = []
    votes = []

    # Fetch candidates and their votes for the chart
    all_candidates = Candidate.objects.all()
    for candidate in all_candidates:
        candidates.append(candidate.name)
        votes.append(candidate.vote_count)

    return render(request, 'chart.html', {'candidates': candidates, 'votes': votes})

def export_results_to_excel(request):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Voting Results"

    # Add the header row
    ws.append(["#", "Position", "Candidate", "Vote Count"])

    # Query all positions and their candidates
    positions = Position.objects.all()
    row_number = 1  # Initialize row number for results
    
    # Iterate over positions and their candidates
    for position in positions:
        candidates = Candidate.objects.filter(position=position)
        
        for idx, candidate in enumerate(candidates, start=1):
            # Write each candidate's position, name, and vote count to the spreadsheet
            row_number += 1
            ws.append([idx, position.name, candidate.name, candidate.vote_count])

    # Adjust the column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter (column_cells[0].column)].width = length + 2

    # Set the HTTP response for the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=voting_results.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


